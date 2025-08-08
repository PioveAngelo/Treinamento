from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

# Carrega uma pasta existente

wb = load_workbook("example.xlsx")

ws = wb.active

# Printa os valores das células selecionadas: A1 e B1, a Saída será "Hello World!"

print(ws['A1'].value)
print(ws['B1'].value)

# Muda a cor da fonte das células selecionadas

ws['A1'].font = Font(bold=True, color="FF0000")
ws['B1'].font = Font(bold=True, color="FFF001")

# Cria um gráfico

values = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E5")

# Salva

wb.save("example.xlsx")