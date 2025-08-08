from openpyxl import Workbook

# Cria uma nova pasta e seleciona a planilha ativa

wb = Workbook()

ws = wb.active

# Adiciona dados na planilha

ws['A1'] = "Hello"
ws['B1'] = "World!"

# Salva a pasta

wb.save("example.xlsx")