import openpyxl
import pandas as pd
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from dateutil.relativedelta import relativedelta
import locale


hoje = datetime.today() # Define a data e hora local
mes_posterior = hoje + relativedelta(months=1) # Pega o próximo mês a partir do mês atual
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8') # Define a data como padrão do Brasil
vencimento = mes_posterior.strftime('%d/%m/%Y') # Define o vencimento como o próximo mês
nome_mes = mes_posterior.strftime('%B') # Pega somente o nome do mês para incluír no e-mail

class Planilha:

    '''
        A classe planilha fica responsável por criar uma planílha com os dados de fornecedores e gera também um exemplo de boleto para enviar como anexo no e-mail. 

        A função cria_planilha() como o nome diz, cria uma planilha a partir de um dicionário e salva o arquivo .xlsx.

        A função gera_boleto() gera um exemplo de boleto muito simples para teste, a partir de um DataFrame da leitura da planilha criada com a função cria_planilha, com os dados lidos da planilha ele pega os dados e inclui em outro arquivo .xlsx em sequência um para cada linha da planilha criada.
    '''

    def cria_planilha(self, dicionario):

        dicionario = {
            'Fornecedor' : ['fornecedor1', 'fornecedor2', 'fornecedor3', 'fornecedor4'],
            'Email' : ['angelopiovezan15@gmail.com', 'fornecedor2@email.com', 'fornecedor3@email.com', 'fornecedor4@email.com'],
            'Fornece' : ['Eletronicos', 'Plasticos', 'Maquinas', 'Internet'],
            'Boleto' : ['boleto1', 'boleto2', 'boleto3', 'boleto4']
        } # Dicionário com as informações de fornecedores fictícios

        nome_planilha = 'dados_fornecedores.xlsx' # Nome que será dado ao arquivo .xlsx
        df = pd.DataFrame(dicionario) # Transforma o dicionário em tabela ordenada onde as keys são os nomes das colunas
        df.to_excel(nome_planilha, index=False) # Transforma em excel (parâmetro 1, parãmetro 2) parâmetro 1: vai o nome que será salvo o arquivo .xlsx, parãmetro 2: estou indicando que não quero que salve os índices das linhas junto ao arquivo

        print(f"Dados salvos com sucesso em {nome_planilha}")


    def gera_boleto(self, df):

        df = pd.read_excel('dados_fornecedores.xlsx') # Atribuo a leitura do arquivo gerado na função cria_planilha() à uma variável chamada 'df' de DataFrame.

        for index, row in df.iterrows(): # Esse loop percorre o DataFrame. df.iterrows() percorre cada linha do DataFrame atribuído ao df = pd.read_excel('dados_fornecedores.xlsx'). O index representa o índice de cada linha e o row pega os dados da linha.
            nome    = row['Fornecedor']
            email   = row['Email']
            produto = row['Fornece']
            boleto  = row['Boleto']
            data    = vencimento
            valor   = '$$$$'


            wb = openpyxl.load_workbook('boleto.xlsx') # Usando o openpyxl eu carrego o 'boleto.xlsx' com o método load_workbook() para alterá-lo e atribuo a variável wb.

            ws = wb.active # Atribuo à variável ws o arquivo que carreguei na variável wb na página em que está ativa quando se abre o arquivo com wb.active, caso tenha outras abas dentro da planilha, pode-se atribuir outra chamada como: wb['nome-da-aba'].

            ws['B1'] = nome # Aqui eu incluo no arquivo em células específicas o que foi atribuido as variáveis anteriormente.
            ws['B2'] = produto
            ws['B3'] = email
            ws['B4'] = valor
            ws['B5'] = data

            wb.save(f"{boleto}.xlsx") # Aqui está sendo salvo o arquivo boleto com o nome que foi percorrido na coluna Boleto do arquivo.

            print(f"Boleto gerado para {nome}, salvo como {boleto}.xlsx")  


class Email:
    '''
        A classe Email é responsável pela função configura_envia_email() que recebe também o DataFrame (df). Como já diz o nome, essa função configura e envia o e-mail com o anexo .xlsx.

        Primeiro atribuímos o DataFrame a variável df, depois setamos as configurações padrão para o servidor SMTP com o gmail. Passamos as informações de autenticação de um e-mail para automatizar o envio do e-mail, passamos o e-mail e a senha configurada para esse processo.

        Após atribuir essas configurações iniciamos um loop que pegará as informações do DataFrame fará um e-mail personalizado para cada indivíduo que tem informações no arquivo dados_fornecedores.xlsx.
    '''
    def configura_envia_email(self, df):

        df = pd.read_excel('dados_fornecedores.xlsx') # Atribuo à variável df a leitura do DataFrame

        smtp_server = "smtp.gmail.com" # Configuração padrão SMTP com gmail
        smtp_port = 587 # Configuração padrão SMTP com gmail
        email_remetente = "piovezanjorgeto@gmail.com" # E-mail do qual será enviado os arquivos
        senha = "cwzg gumc ifqx kdls" # Senha de acesso para o envio dos e-mails

        for index, row in df.iterrows(): # Esse loop percorre o DataFrame. df.iterrows() percorre cada linha do DataFrame atribuído ao df = pd.read_excel('dados_fornecedores.xlsx'). O index representa o índice de cada linha e o row pega os dados da linha.
            
            nome    = row['Fornecedor'] # Atribuo à variáveis as informações coletadas para o envio do e-mail
            destino = row['Email']
            boleto  = row['Boleto']
            arquivo   = f'{boleto}.xlsx'

            msg = MIMEMultipart() # Atribuo a variável msg a classe MIMEMultipart() da biblioteca email.mime.multipart que permite criar e-mails com múltiplas partes como: Texto(corpo do e-mail), HtML, anexos(.xlsx, PDFs, etc), etc. Então msg = MIMEMultipart() está criando um objeto de e-mail vazio, esse objeto pode ser preenchido com remetente, destinatário, assunto, corpo do e-mail e anexos como vemos a seguir.
            msg['From']    = email_remetente
            msg['To']      = destino
            msg['Subject'] = f'Boleto com vencimento em {mes_posterior}'
            corpo_email    = f'Olá {nome}, segue o boleto com vencimento para {vencimento}.\n\nFavor pagar até a data de vencimento para não haver juros.\n\nAtenciosamente Britânia Eletrodomésticos!'
            msg.attach(MIMEText(corpo_email, "plain")) # Aqui estmaos adicionando o corpo do e-mail no objeto msg e com o "plain" dizemos que será somente texto simples, sem html. attach adiciona ao e-mail completo.

            with open(arquivo, "rb") as anexo: # with é uma função que abre e fecha de forma segura, pois mesmo que aconteça algum erro ele fechará o que foi aberto. open(arquivo, "rb") abre o arquivo f'{boleto}.xlsx' atribuido à variável aquivo anteriormente de modo binário para leitura.

                anexo_bin = MIMEBase("application", "octet-stream") # Cria uma estrutura MIME genérica para arquivos binário.
                anexo_bin.set_payload(anexo.read()) # Lê o conteúdo do arquivo e define como "corpo" do anexo.
                encoders.encode_base64(anexo_bin) # Codifica em base64 para um envio seguro por e-mail.
                anexo_bin.add_header("Content-Disposition", f"atachment; filename={arquivo}") # Adiciona um cabeçalho que diz ao cliente de e-mail que o que está indo é um anexo, com o nome do arquivo
                msg.attach(anexo_bin) # Adiciona esse anexo ao e-mail que está sendo enviado.

            with smtplib.SMTP(smtp_server, smtp_port) as server: # Aqui também usamos o with para uma abertura de conexão segura, que será fechada assim que o processo for concluído mesmo que ocorra algum erro. fazemos a conexão com o servidor SMTP com smtplib.SMTP(smtp_server, smtp_port) e passamos os parâmetros setados anteriormente.
                server.starttls() # Aqui iniciamos o TLS para uma comunicação segura com criptografia.
                server.login(email_remetente, senha) # aqui fazemos o login para o envio do e-mail com os dados de autenticação setados anteriormente.
                server.sendmail(email_remetente, destino, msg.as_string()) # Aqui enviamos o e-mail com os parâmetros setados anteriormente e convertendo o objeto msg para string com msg.as_string()
                
            print(f"E-mail enviado para {destino} com o boleto {arquivo}")

    