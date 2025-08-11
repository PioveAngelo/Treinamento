import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel('dados_clientes.xlsx')

def gerar_nota():

    for index, row in df.iterrows():
        nome = row['Nome']
        email = row['Email']
        produto = row['Produto']
        quantidade = row['Quantidade']
        valor_unitario = row['Valor Unitário']
        valor_total = quantidade * valor_unitario

        wb = load_workbook('modelo_nota.xlsx')
        ws = wb.active

        ws['B2'] = nome
        ws['B3'] = email
        ws['B4'] = produto
        ws['B5'] = quantidade
        ws['B6'] = valor_unitario
        ws['B7'] = valor_total

        wb.save(f'nota_{nome}.xlsx')
        print(f"Nota gerada para {nome} e salva como 'nota_{nome}.xlsx'")
    
if __name__ == "__main__":
    gerar_nota()
    print("Todas as notas foram geradas com sucesso.")